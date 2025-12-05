# XE.py
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter.font import Font
import mysql.connector
import openpyxl
from openpyxl.styles import Font, Alignment
# ===================== THIáº¾T Láº¬P MÃ€U Sáº®C ======================
BG_COLOR = "#f0f8ff"      # MÃ u ná»n tá»•ng thá»ƒ (Alice Blue)
HEADER_COLOR = "#4682b4"  # MÃ u tiÃªu Ä‘á» (Steel Blue)
LABEL_COLOR = "#333333"   # MÃ u chá»¯ Label
BUTTON_BG = "#5cb85c"     # MÃ u ná»n nÃºt (Green - Success)
BUTTON_FG = "white"       # MÃ u chá»¯ nÃºt
FRAME_BG = "#e9f5ff"      # MÃ u ná»n Frame thÃ´ng tin
ACCENT_COLOR = "#dc3545"  # MÃ u nháº¥n (Error/Delete)

# ===================== Káº¾T Ná»I DATABASE ======================
def connect_db():
    try:
        return mysql.connector.connect(
            host="localhost",
            user="root",
            password="123456",
            database="qli_chxm",
        )
    except mysql.connector.Error as err:
        messagebox.showerror("Lá»—i DB", f"KhÃ´ng thá»ƒ káº¿t ná»‘i CSDL: {err}")
        return None

# ===================== HÃ€M CÄ‚N GIá»®A ==========================
def center_window(win, w=950, h=650):
    """Äáº·t cá»­a sá»• vÃ o giá»¯a mÃ n hÃ¬nh."""
    ws = win.winfo_screenwidth()
    hs = win.winfo_screenheight()
    x = (ws // 2) - (w // 2)
    y = (hs // 2) - (h // 2)
    win.geometry(f"{w}x{h}+{x}+{y}")

# ===================== HÃ€M HIá»‚N THá»Š FORM XE ===================
def show(root):
    win = tk.Toplevel(root)
    win.title("Quáº£n lÃ½ XE")
    center_window(win, w=950, h=650)
    win.configure(bg=BG_COLOR)
    win.transient(root) # Äáº·t form con luÃ´n náº±m trÃªn form cha

    # ===== TIÃŠU Äá»€ =====
    tk.Label(win, text="QUáº¢N LÃ THÃ”NG TIN XE", 
             font=("Arial", 18, "bold"), 
             bg=HEADER_COLOR, fg="white", 
             pady=10).pack(fill=tk.X)

    # ===== FRAME THÃ”NG TIN XE (2 cá»™t nháº­p liá»‡u) =====
    frame_info = tk.LabelFrame(win, text="ğŸ“ Chi tiáº¿t sáº£n pháº©m", 
                               font=("Times New Roman", 13, "bold"), 
                               padx=20, pady=15, bg=FRAME_BG, fg=HEADER_COLOR)
    frame_info.pack(side="bottom",padx=20, pady=15, fill="x")

    # Äáº·t trá»ng sá»‘ cho cá»™t Ä‘á»ƒ Entry má»Ÿ rá»™ng Ä‘áº¹p hÆ¡n
    frame_info.grid_columnconfigure(1, weight=1)
    frame_info.grid_columnconfigure(3, weight=1)

    label_font = ("Times New Roman", 12)
    entry_font = ("Times New Roman", 12)
    
    # Dictionary lÆ°u Entry
    entries = {}
    
    # Danh sÃ¡ch cÃ¡c trÆ°á»ng cáº§n nháº­p liá»‡u
    fields = [
        ("MÃ£ SP", 0, 0), ("TÃªn SP", 1, 0), 
        ("Loáº¡i xe", 2, 0), ("HÃ£ng SX", 0, 2), 
        ("GiÃ¡ (VND)", 1, 2), ("Sá»‘ lÆ°á»£ng (Kho)", 2, 2)
    ]
    
    for text, row, col in fields:
        tk.Label(frame_info, text=text, font=label_font, bg=FRAME_BG, fg=LABEL_COLOR).grid(
            row=row, column=col, sticky="w", pady=5, padx=(0, 10)) # sticky="w" cÄƒn trÃ¡i Label

        entry = tk.Entry(frame_info, font=entry_font, borderwidth=1, relief="solid")
        # Äáº·t Entry vÃ o cá»™t tiáº¿p theo vÃ  má»Ÿ rá»™ng theo chiá»u ngang (sticky="ew")
        entry.grid(row=row, column=col+1, padx=(0, 20), pady=5, sticky="ew")
        entries[text] = entry
        
    # GÃ¡n biáº¿n Entry cho dá»… sá»­ dá»¥ng
    entry_masp = entries["MÃ£ SP"]
    entry_tensp = entries["TÃªn SP"]
    entry_loaixe = entries["Loáº¡i xe"]
    entry_hangsx = entries["HÃ£ng SX"]
    entry_gia = entries["GiÃ¡ (VND)"]
    entry_soluong = entries["Sá»‘ lÆ°á»£ng (Kho)"]

    # ===== FRAME NÃšT CHá»¨C NÄ‚NG =====
     # ===== FRAME NÃšT CHá»¨C NÄ‚NG =====


    btn_font = ("Times New Roman", 13, "bold")
    

    def xuat_excel():
        if not tree.get_children():
            messagebox.showwarning("KhÃ´ng cÃ³ dá»¯ liá»‡u", "KhÃ´ng cÃ³ dá»¯ liá»‡u Ä‘á»ƒ xuáº¥t!")
            return
    
        try:
        # Táº¡o workbook má»›i
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Danh sÃ¡ch Xe"

        # Ghi tiÃªu Ä‘á» cá»™t
            headers = ["MÃ£ SP", "TÃªn SP", "Loáº¡i Xe", "HÃ£ng SX", "GiÃ¡ (VND)", "Sá»‘ lÆ°á»£ng (Kho)"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

        # Ghi dá»¯ liá»‡u tá»« Treeview
            for row_num, item in enumerate(tree.get_children(), 2):
                values = tree.item(item)["values"]
                for col_num, value in enumerate(values, 1):
                    ws.cell(row=row_num, column=col_num, value=value)

        # Tá»± Ä‘á»™ng cÄƒn chá»‰nh Ä‘á»™ rá»™ng cá»™t
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # LÆ°u file
            file_path = "Danh_sach_xe.xlsx"
            wb.save(file_path)
            messagebox.showinfo("Xuáº¥t Excel thÃ nh cÃ´ng", f"ÄÃ£ xuáº¥t dá»¯ liá»‡u ra file {file_path}")
    
        except Exception as e:
            messagebox.showerror("Lá»—i", f"KhÃ´ng thá»ƒ xuáº¥t Excel: {e}")

    # ======================= HÃ€M Xá»¬ LÃ =========================
    
    def load_data():
        tree.delete(*tree.get_children())
        conn = connect_db()
        if not conn: return
        try:
            cur = conn.cursor()
            cur.execute("SELECT MaSP, TenSP, LoaiXe, Hangsx, Gia, SoLuong FROM XE")
            for row in cur.fetchall():
                # Äá»‹nh dáº¡ng giÃ¡ tiá»n (VÃ­ dá»¥: 1000000 -> 1,000,000)
                formatted_gia = f"{row[4]:,.0f}" if row[4] is not None else "N/A"
                # Táº¡o hÃ ng má»›i vá»›i giÃ¡ Ä‘Ã£ Ä‘á»‹nh dáº¡ng
                tree.insert("", "end", values=(row[0], row[1], row[2], row[3], formatted_gia, row[5]))
        except Exception as e:
            messagebox.showerror("Lá»—i DB", str(e))
        finally:
            if conn and conn.is_connected():
                conn.close()

    def clear_input():
        for e in entries.values():
            e.delete(0, tk.END)

    def validate_input(is_update=False):
        """HÃ m kiá»ƒm tra dá»¯ liá»‡u Ä‘áº§u vÃ o."""
        data = {
            'masp': entry_masp.get().strip(),
            'tensp': entry_tensp.get().strip(),
            'loaixe': entry_loaixe.get().strip(),
            'hangsx': entry_hangsx.get().strip(),
        }
        
        if not data['masp'] or not data['tensp']:
            messagebox.showwarning("Thiáº¿u dá»¯ liá»‡u", "MÃ£ SP vÃ  TÃªn SP báº¯t buá»™c!")
            return None

        try:
            data['gia'] = float(entry_gia.get().strip().replace(',', ''))
            data['sl'] = int(entry_soluong.get().strip())
        except ValueError:
            messagebox.showwarning("Lá»—i dá»¯ liá»‡u", "GiÃ¡ pháº£i lÃ  sá»‘, Sá»‘ lÆ°á»£ng pháº£i lÃ  sá»‘ nguyÃªn")
            return None
            
        if data['gia'] < 0 or data['sl'] < 0:
            messagebox.showwarning("Lá»—i dá»¯ liá»‡u", "GiÃ¡ vÃ  Sá»‘ lÆ°á»£ng pháº£i lá»›n hÆ¡n hoáº·c báº±ng 0.")
            return None

        return data
#thenh xoa sua
    def them():
        data = validate_input()
        if not data: return

        conn = connect_db()
        if not conn: return
        try:
            cur = conn.cursor()
            cur.execute("INSERT INTO XE (MaSP, TenSP, LoaiXe, Hangsx, Gia, SoLuong) VALUES (%s,%s,%s,%s,%s,%s)",
                        (data['masp'], data['tensp'], data['loaixe'], data['hangsx'], data['gia'], data['sl']))
            conn.commit()
            messagebox.showinfo("ThÃ nh cÃ´ng", f"ÄÃ£ thÃªm xe {data['tensp']} ({data['masp']})")
            load_data()
            clear_input()
        except mysql.connector.errors.IntegrityError:
            messagebox.showerror("Lá»—i", "MÃ£ SP Ä‘Ã£ tá»“n táº¡i!")
        except Exception as e:
            messagebox.showerror("Lá»—i", str(e))
        finally:
            if conn and conn.is_connected(): conn.close()

    def sua():
        sel = tree.selection()
        if not sel:
            messagebox.showwarning("Chá»n dá»¯ liá»‡u", "HÃ£y chá»n xe Ä‘á»ƒ sá»­a")
            return

        item_id = sel[0]
        ma_cu = tree.item(item_id)['values'][0] 
        data = validate_input(is_update=True)
        if not data: return

        if data['masp'] != ma_cu:
            messagebox.showwarning("Cáº£nh bÃ¡o", "KhÃ´ng Ä‘Æ°á»£c thay Ä‘á»•i MÃ£ SP khi Sá»­a. HÃ£y nháº¥n Há»§y vÃ  chá»n láº¡i.")
            return

        conn = connect_db()
        if not conn: return
        try:
            cur = conn.cursor()
            cur.execute("UPDATE XE SET TenSP=%s, LoaiXe=%s, Hangsx=%s, Gia=%s, SoLuong=%s WHERE MaSP=%s",
                        (data['tensp'], data['loaixe'], data['hangsx'], data['gia'], data['sl'], ma_cu))
            conn.commit()
            
            # Cáº­p nháº­t trá»±c tiáº¿p trÃªn Treeview vá»›i giÃ¡ Ä‘Ã£ Ä‘á»‹nh dáº¡ng
            formatted_gia = f"{data['gia']:,.0f}"
            tree.item(item_id, values=(ma_cu, data['tensp'], data['loaixe'], data['hangsx'], formatted_gia, data['sl']))
            
            messagebox.showinfo("ThÃ nh cÃ´ng", f"ÄÃ£ cáº­p nháº­t thÃ´ng tin xe {ma_cu}")
            clear_input()
        except Exception as e:
            messagebox.showerror("Lá»—i", str(e))
        finally:
            if conn and conn.is_connected(): conn.close()

    def xoa():
        sel = tree.selection()
        if not sel:
            messagebox.showwarning("Chá»n dá»¯ liá»‡u", "HÃ£y chá»n xe Ä‘á»ƒ xÃ³a")
            return

        ma = tree.item(sel[0])['values'][0]
        if messagebox.askyesno("XÃ¡c nháº­n", f"Báº¡n cháº¯c cháº¯n muá»‘n xÃ³a xe {ma}?"):
            conn = connect_db()
            if not conn: return
            try:
                cur = conn.cursor()
                cur.execute("DELETE FROM XE WHERE MaSP=%s", (ma,))
                conn.commit()
                load_data()
                clear_input()
            except Exception as e:
                messagebox.showerror("Lá»—i", str(e))
            finally:
                if conn and conn.is_connected(): conn.close()

    # ===== CHá»ŒN DÃ’NG TREEVIEW =====
    def on_tree_select(event):
        sel = tree.selection()
        if not sel: return
        
        values = tree.item(sel[0])["values"]
        clear_input() 
        
        # Chuyá»ƒn Ä‘á»•i giÃ¡ tá»« chuá»—i Ä‘á»‹nh dáº¡ng (cÃ³ dáº¥u pháº©y) vá» dáº¡ng sá»‘ (khÃ´ng dáº¥u pháº©y) trÆ°á»›c khi hiá»ƒn thá»‹
        gia_str = str(values[4]).replace(',', '') 
        
        entry_masp.insert(0, values[0])
        entry_tensp.insert(0, values[1])
        entry_loaixe.insert(0, values[2])
        entry_hangsx.insert(0, values[3])
        entry_gia.insert(0, gia_str)
        entry_soluong.insert(0, values[5])


    # ===== TREEVIEW HIá»‚N THá»Š Dá»® LIá»†U =====
    cols = ("MaSP", "TenSP", "LoaiXe", "Hangsx", "Gia", "SoLuong")
    
    # Táº¡o Scrollbar dá»c
    vsb = ttk.Scrollbar(win, orient="vertical")
    
    style = ttk.Style()
    style.theme_use('clam')
    style.configure("Treeview.Heading", font=('Times New Roman', 12, 'bold'), background=HEADER_COLOR, foreground="white")
    style.configure("Treeview", font=('Times New Roman', 11), rowheight=25)
    
    tree = ttk.Treeview(win, columns=cols, show="headings", yscrollcommand=vsb.set)
    vsb.config(command=tree.yview)

    # ÄÃ³ng gÃ³i scrollbar vÃ  treeview (sá»­ dá»¥ng expand=True Ä‘á»ƒ nÃ³ chiáº¿m khoáº£ng trá»‘ng lá»›n nháº¥t)
    vsb.pack(side='right', fill='y', padx=(0, 20))
    tree.pack(fill="both", expand=True, padx=(20, 0), pady=(10, 5)) # Giáº£m pady dÆ°á»›i

    for c in cols:
        tree.heading(c, text=c)
        tree.column(c, anchor="center", width=120)
    
    tree.column("Gia", anchor="e", width=150)
    tree.column("TenSP", anchor="w", width=180)

    for c in cols:
        tree.heading(c, text=c)
        tree.column(c, anchor="center", width=120)
    
    # CÄƒn chá»‰nh cá»™t GiÃ¡ vÃ  TÃªn SP rá»™ng hÆ¡n
    tree.column("Gia", anchor="e", width=150)
    tree.column("TenSP", anchor="w", width=180)
    
    tree.bind("<<TreeviewSelect>>", on_tree_select)
    load_data()
    
    # ===== FRAME NÃšT CHá»¨C NÄ‚NG (ÄÆ°a xuá»‘ng dÆ°á»›i TreeView) =====
    frame_btn = tk.Frame(win, bg=BG_COLOR)
    frame_btn.pack(pady=(10, 15))

    btn_font = ("Times New Roman", 13, "bold")

    tk.Button(frame_btn, text="â• ThÃªm", width=12, font=btn_font, command=them, bg=BUTTON_BG, fg=BUTTON_FG).pack(side="left", padx=7)
    tk.Button(frame_btn, text="âœï¸ Sá»­a", width=12, font=btn_font, command=sua, bg="#ffc107", fg=LABEL_COLOR).pack(side="left", padx=7)
    tk.Button(frame_btn, text="ğŸ—‘ï¸ XÃ³a", width=12, font=btn_font, command=xoa, bg=ACCENT_COLOR, fg=BUTTON_FG).pack(side="left", padx=7)
    tk.Button(frame_btn, text="ğŸ”„ Há»§y", width=12, font=btn_font, command=clear_input, bg="#6c757d", fg=BUTTON_FG).pack(side="left", padx=7)
    tk.Button(frame_btn, text="ğŸšª ThoÃ¡t", width=12, font=btn_font, command=win.destroy, bg="#343a40", fg=BUTTON_FG).pack(side="left", padx=7)
    tk.Button(frame_btn, text="ğŸ“„ Xuáº¥t Excel", width=12, font=btn_font, command=xuat_excel, bg="#17a2b8", fg=BUTTON_FG).pack(side="left", padx=7)

    # KhÃ´ng nÃªn gá»i win.mainloop() á»Ÿ Ä‘Ã¢y, Ä‘á»ƒ root.mainloop() trong main.py quáº£n lÃ½
    # win.mainloop() 
    return win